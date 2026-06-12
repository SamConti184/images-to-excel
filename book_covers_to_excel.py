#!/usr/bin/env python3
"""
Book Cover Metadata Extractor
Reads Italian book cover images from a folder, extracts metadata using GPT vision,
and saves results to an Excel file.

Usage:
    book_covers_to_excel.exe <folder_path> <openai_api_key>
    book_covers_to_excel.exe                          (will prompt for inputs)
"""

import os
import sys
import base64
import json
import time
from pathlib import Path

try:
    from openai import OpenAI, RateLimitError, APIError
except ImportError:
    print("ERROR: openai package is not installed. Cannot continue.")
    input("Press Enter to exit...")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is not installed. Cannot continue.")
    input("Press Enter to exit...")
    sys.exit(1)

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp'}
MODEL = "gpt-5.5"

SYSTEM_PROMPT = """Sei un esperto di catalogazione libraria italiana.
Analizza la copertina del libro nell'immagine ed estrai le seguenti informazioni.
Rispondi SOLO con un oggetto JSON valido, senza testo aggiuntivo, senza markdown, senza backtick.
Il JSON deve avere esattamente queste chiavi:
{
  "autore": "cognome e nome dell'autore, o stringa vuota se non visibile",
  "titolo": "titolo completo del libro, o stringa vuota se non visibile",
  "luogo_di_stampa": "città di stampa/pubblicazione, o stringa vuota se non visibile",
  "editore": "nome dell'editore, o stringa vuota se non visibile",
  "anno": "anno di pubblicazione come stringa, o stringa vuota se non visibile"
}
Se un'informazione non è visibile sulla copertina, lascia il campo come stringa vuota.
Non inventare informazioni non presenti nell'immagine."""


def image_to_base64(image_path):
    """Read image file and return (base64_string, media_type)."""
    ext = Path(image_path).suffix.lower()
    media_type_map = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.webp': 'image/webp',
    }
    media_type = media_type_map.get(ext, 'image/jpeg')
    with open(image_path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode('utf-8')
    return b64, media_type


def call_openai_vision(client, image_path, max_retries=5):
    """Call OpenAI vision API with exponential backoff on rate limit errors."""
    b64, media_type = image_to_base64(image_path)

    for attempt in range(max_retries):
        try:
            response = client.responses.create(
                model=MODEL,
                input=[
                    {
                        "role": "system",
                        "content": SYSTEM_PROMPT
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "input_text",
                                "text": "Estrai i metadati da questa copertina."
                            },
                            {
                                "type": "input_image",
                                "image_url": f"data:{media_type};base64,{b64}",
                                "detail": "high"
                            }
                        ]
                    }
                ]
            )
            content = response.output_text.strip()
            # Clean possible markdown fences just in case
            if content.startswith('```'):
                content = content.split('```')[1]
                if content.startswith('json'):
                    content = content[4:]
            return json.loads(content.strip())

        except RateLimitError:
            wait = 2 ** attempt  # 1s, 2s, 4s, 8s, 16s
            print(f"\n  [rate limited, waiting {wait}s...]", end='', flush=True)
            time.sleep(wait)

        except APIError as e:
            raise RuntimeError(f"OpenAI API error: {e}")

    raise RuntimeError(f"Failed after {max_retries} retries due to rate limiting.")


def get_image_files(folder):
    files = []
    for f in sorted(os.listdir(folder)):
        if Path(f).suffix.lower() in IMAGE_EXTENSIONS:
            files.append(os.path.join(folder, f))
    return files


def save_to_excel(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libri"

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='2E4057')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'File', 'Autore', 'Titolo', 'Luogo di Stampa', 'Editore', 'Anno', 'Note']
    col_widths = [5, 30, 25, 40, 20, 25, 8, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    row_fill_even = PatternFill('solid', start_color='F5F7FA')
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)

    for row_idx, entry in enumerate(results, 2):
        fill = row_fill_even if row_idx % 2 == 0 else PatternFill()
        values = [
            row_idx - 1,
            entry.get('filename', ''),
            entry.get('autore', ''),
            entry.get('titolo', ''),
            entry.get('luogo_di_stampa', ''),
            entry.get('editore', ''),
            entry.get('anno', ''),
            entry.get('note', ''),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if fill.fill_type:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = 'A2'
    wb.save(output_path)


def main():
    print("=" * 60)
    print("  Book Cover Metadata Extractor (Italian)")
    print("  Powered by OpenAI GPT-5.5 Vision")
    print("=" * 60)
    print()

    if len(sys.argv) >= 2:
        folder = sys.argv[1].strip().strip('"')
    else:
        folder = input("Enter the folder path containing book cover images:\n> ").strip().strip('"')

    if not os.path.isdir(folder):
        print(f"\nERROR: '{folder}' is not a valid folder.")
        input("Press Enter to exit...")
        sys.exit(1)

    if len(sys.argv) >= 3:
        api_key = sys.argv[2].strip()
    else:
        api_key = input("\nEnter your OpenAI API key:\n> ").strip()

    if not api_key.startswith('sk-'):
        print("\nWARNING: API key doesn't look right (should start with 'sk-'). Continuing anyway...")

    client = OpenAI(api_key=api_key)

    image_files = get_image_files(folder)
    if not image_files:
        print(f"\nNo image files found in '{folder}'.")
        input("Press Enter to exit...")
        sys.exit(0)

    print(f"\nFound {len(image_files)} image(s) to process.")
    confirm = input("Proceed? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Aborted.")
        input("Press Enter to exit...")
        sys.exit(0)

    print()
    results = []
    errors = 0

    for i, image_path in enumerate(image_files, 1):
        filename = os.path.basename(image_path)
        print(f"[{i:3d}/{len(image_files)}] Processing: {filename}", end='', flush=True)

        entry = {'filename': filename}

        try:
            data = call_openai_vision(client, image_path)
            entry.update({
                'autore': data.get('autore', ''),
                'titolo': data.get('titolo', ''),
                'luogo_di_stampa': data.get('luogo_di_stampa', ''),
                'editore': data.get('editore', ''),
                'anno': data.get('anno', ''),
                'note': ''
            })
            print(f"  ✓  {data.get('titolo', '(no title)')[:50]}")
        except Exception as e:
            entry.update({
                'autore': '', 'titolo': '', 'luogo_di_stampa': '',
                'editore': '', 'anno': '', 'note': f'ERROR: {str(e)[:80]}'
            })
            print(f"  ✗  ERROR: {str(e)[:60]}")
            errors += 1

        results.append(entry)

    output_path = os.path.join(folder, "libri_catalogo.xlsx")
    print(f"\nSaving results to Excel...")
    save_to_excel(results, output_path)

    print(f"\n{'=' * 60}")
    print(f"  Done! Processed {len(image_files)} image(s), {errors} error(s).")
    print(f"  Saved to: {output_path}")
    print(f"{'=' * 60}")
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Unexpected error: {e}")
        input("Press Enter to exit...")